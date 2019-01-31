#!/home/eliot/anaconda2/bin/python2.7

#delete comment for rerun before use
import os,sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

import plotly as py
import plotly.figure_factory as ff
from plotly.graph_objs import *
import scipy 

import argparse
parser = argparse.ArgumentParser()                                               

parser.add_argument("--file", "-f", type=str, required=True)
parser.add_argument("--method", "-m", type=str, default = "bmas")
parser.add_argument("--total_marker", "-t", type=int, default = 0)
parser.add_argument("--generation", "-g", type=str, required=True)

args = parser.parse_args()
##################################################################################################
############################################## In/Out data #######################################
#################################################################################################
# Read raw data
wb = pd.ExcelFile(args.file)

rawData = wb.parse('Sheet1')
wb.close()


#w_retestMarker = open('retestMarker','w')

wb = load_workbook(args.file)
sheet_AB = wb.create_sheet('A-B')
sheet_Result = wb.create_sheet('Results')
sheet_Err = wb.create_sheet('Error')

method = args.method
totalMarker = args.total_marker
gen = args.generation

##################################################################################################
######################################### Criteria setting #######################################
##################################################################################################

## A-B convert
voteCutoff = 80		# If the proportions of 2 dyes differ < 80%, the result will be counted as blank.
blankAcc = 5		# If blank > 5%, cut out or retest

## Sample selection for back crossing
ref_perc = 10/100
sp_perc = 20/100

## Minimum espectation of %BG
minEspect = {'f1':50 , 'bc1': 75.00, 'bc2': 87.50, 'bc3': 93.75, 'bc4': 96.88, 'bc5': 98.44, 'bc6': 99.22, 'bc7':99.61, 'bc8': 99.80, 'bc9': 99.90, 'bc10': 99.95, 'bc11':99.98, 'bc12': 99.99}
if gen not in minEspect.keys():
	minEspect = 100
else:
	minEspect = minEspect[gen]

##################################################################################################
######################################### BMAS Manegement ########################################
##################################################################################################
if method == 'bmas' :
	# Transpose raw data
	#print (rawData)
	rawData = rawData.T
	#print (rawData)

	##################### Define references and sample	#######################
	fiRef_row = 0
	laRef_row = 0
	fiSp_row = 0
	laSp_row = 0

	listRawData = list(rawData.index)
	i=1
     	#print (listRawData[i])
	if "RP-B" not in str(listRawData[i]):	# Detect whether data start from sample or not?
		#print(listRawData)
		fiSp_row = 1
		while "RP-B" not in str(listRawData[i]):
			i+=1
		laSp_row = i
		fiRef_row = i
		laRef_row = len(listRawData)
	else:	
		if "RP-B" in listRawData[i]:	# Detect whether data start from sample or not?
			fiRef_row = 1
			while "RP-B" in str(listRawData[i]):
				i+=1
			laRef_row = i
			fiSp_row = i+1
			laSp_row = len(listRawData)
		else:
			print('Please check your format!!!!!')
			
	######## Vote V-F for each marker and keep voted dye in dictionary >> dict{'marker1':'F','marker2':'V'} ########
	errMarker = []
	voteDye = {}

	sheet_Err.cell(('A1')).value='Marker(s) with blank and A > 5%'
	errMarker1_line = 2
	sheet_Err.cell(('C1')).value='Marker(s) returned equal V-F'
	errMarker2_line = 2
	#sheet_Result.cell(row =1 , column =1 ,value='Markers for the next generation')
	retestMarker_line =2

	for j in rawData[fiRef_row:laRef_row]:	# each j will show all dyes for each marker
		voteDye[j] = ''   #use column number as keys
		countF = list(rawData[j])[fiRef_row:laRef_row].count('F')
		countV = list(rawData[j])[fiRef_row:laRef_row].count('V')
		#print (j,countF,countV)
		if countF > countV :
			voteDye[j] = 'F'
		else:
			if countV > countF:
				voteDye[j] = 'V'
			else:
				voteDye[j] = 'na'
				#errMarker.append(j)   #comment for rerun
				sheet_Err.cell(row = errMarker2_line, column = 3, value = rawData[j][0])
				errMarker2_line+=1
				print ('same dye: ',j,countF,countV)
		#print (list(rawData[j])[fiRef_row:laRef_row])
		#print (voteDye[j])

	######################################### Write retest/error marker ###########################################
	retestMarker=[]

	for i in rawData:
		if (list(rawData[i]).count("-")+list(rawData[i]).count("A"))*100/(len(list(rawData[i]))-1) > blankAcc:	#if any markers showed blank+A results > 5%, cut out or retest
			#errMarker.append(i)    #comment for rerun
			sheet_Err.cell(row = errMarker1_line, column = 1, value = rawData[i][0])
			errMarker1_line+=1
			#print ('blank>5: ',i, list(rawData[i]).count("-"), list(rawData[i]).count("A"), len(list(rawData[i])), (list(rawData[i]).count("-")+list(rawData[i]).count("A"))*100/len(list(rawData[i])))
		else:
			if "H" in list(rawData[i]) or "-" in list(rawData[i]):
				#sheet_Result.cell(row =retestMarker_line , column =1 ,value=rawData[i][0])
				retestMarker_line+=1
	if gen != 'bc1':	#if BC1, markers with blank > 5% and varied dyes will not be calculated
		errMarker = []
	#print (errMarker)
		
	###################### A-B convert and calcualtion, and keep data in excel ##############################
	rowNum = 1
	data=[]
	text=[]
	plantMoreExpect=0
	#print (voteDye)
	if gen == 'bc1':
		#print (totalMarker)
		totalMarker -= len(errMarker)

	for i in list(rawData.index):	# do for each sample; i = individual plant
		colNum = 2
		countH = 0
		countB = 0
		countA = 0
		countBlank = 0
		#print(i)
		#print (rawData.loc[i])
		sheet_AB.cell(row=rowNum, column=1, value=i)		#sample name
		countCol=0

		for j in rawData.loc[i]:		#j = dye in each marker
			#print countCol
			if countCol not in errMarker:	#
				#print(j,'\t',voteDye[colNum-2],'\t', str(j) != str(voteDye[colNum-2].upper()))
				#print('sample: ',j)
				if j.replace(' ','') == (voteDye[countCol]).upper() and (j in ['V','v','F','f']):	#F1 = PR and F1 is homo >> B
					countB += 1
					if (voteDye[countCol]).isupper():
						sheet_AB.cell(row=rowNum, column=colNum, value='B')
						colNum +=1
					else:
						sheet_AB.cell(row=rowNum, column=colNum, value='b')
				else:
					if j.replace(' ','') == '-':
						countBlank += 1
						sheet_AB.cell(row=rowNum, column=colNum, value= j)
						colNum +=1			
					else:
						if j.replace(' ','') != (voteDye[countCol]).upper() :
							if (j in ['V','v','F','f']):	#F1 != PR and F1 is homo >> A
								countA+=1
								if (voteDye[countCol]).isupper():
									sheet_AB.cell(row=rowNum, column=colNum, value='A')
									colNum +=1
									#print(rowNum)
									#print(colNum)
								else:
									sheet_AB.cell(row=rowNum, column=colNum, value='a')	
									#print(j,'\t',voteDye[colNum-2])	
							else:
								if voteDye[countCol] == 'na':				#PR was voted blank >> -
									sheet_AB.cell(row=rowNum, column=colNum, value= '-')	
									colNum +=1
								else:
									#print('endcode: ',j)
									sheet_AB.cell(row=rowNum, column=colNum, value= j)			
									if j.replace(' ','') == 'H':
										countH +=1	
										colNum +=1
              		"""else:
				countB+=1
				sheet_AB.cell(row=rowNum, column=colNum, value= "-")"""

			if str(i).lower() == 'marker_code' and (countCol not in errMarker):
				sheet_AB.cell(row=rowNum, column=colNum, value= j)
				#print (j,colNum)

				colNum +=1
			countCol +=1

		fixMarker=totalMarker - len(list(rawData))

		## P'Fon formular
		if (gen=='bc1'):
			#totalMarker = len(list(rawData))-len(errMarker)+fixMarker-countBlank
			#percRPG = (float(countB)*2+countH+fixMarker*2)*100/(totalMarker*2)
			percRPG = (1-((float(countA)+(countH/2)+countBlank)/totalMarker))*100
		else:
			## K'Eliot formular
			percRPG = (1-((float(countA)+(countH)+countBlank)/totalMarker))*100
		if "RP-B" not in str(i) and "code" not in str(i):
			sheet_AB.cell(row=rowNum, column=colNum, value= '%.2f'%percRPG)
			if percRPG >= minEspect:
				plantMoreExpect+=1

  		#print(i,'\t',countB,'\t',countA,'\t',countH,'\t',countBlank,'\t',totalMarker,'\t',percRPG)
		rowNum +=1

		percRPG = round((percRPG),2)
		data.append(percRPG)
		text.append(i)
		#print (i,'\t',percRPG)
			#print (data)
	#################################### Collect top rank ###############################################
	refNum_collect = int((laRef_row - fiRef_row +1)*ref_perc)
	spNum_collect = int((laSp_row - fiSp_row +1)*sp_perc)
	rf_bgScore = pd.DataFrame({'sample' : text[fiRef_row:laRef_row] , 'value': data[fiRef_row:laRef_row] })
	rf_sel = rf_bgScore.nlargest(refNum_collect,'value')
	sp_bgScore = pd.DataFrame({'sample' : text[fiSp_row:laSp_row] , 'value' : data[fiSp_row:laSp_row] })
	sp_sel = sp_bgScore.nlargest(3,'value')

	################################# Create normal distribution ###########################
	#x1 = data
	hist_data = data[fiSp_row:laSp_row+1]
	group_labels = ['Group 1']
	colorscale = [[0,'pink'],[1,'red']]
	color = [0]*((100-int(min(hist_data)))+1)	#for create histogram
	label = [""]*((100-int(min(hist_data)))+1)
	#print (hist_data)
	#print(int(max(hist_data)-min(hist_data))+1)
	#print (len(color))
	#print (sp_sel)
	sheet_Result.cell(row = 1, column =1 , value = 'The plants contained high %RPG')

	html_candidate="""<html><p>Plant candidate(s) for back crossing in next generation</p>
		<table border="1">
		<tr><th>Rank</th><th>Plant ID</th><th>%RPG</th><th>Number of markers for next generation</th></tr>
	"""+'\n'

	k=0
	for i in list(sp_sel.index):
		sp=list(sp_sel.loc[[i]]['sample'])
		val=list(sp_sel.loc[[i]].value)
		if val[0] >= minEspect:
			#print (val)
			indexMark = int(val[0])
			#print (indexMark)
			color[indexMark-int(min(hist_data))] = 50
			#print (sp)

		j=4

		sheet_Result.cell(row = 2, column =1+(4*k) ,value = '#'+str(k+1))
		sheet_Result.cell(row = 3, column =1+(4*k) ,value = 'Plant_id')
		sheet_Result.cell(row = 3, column =2+(4*k) ,value = '%RPG')
		sheet_Result.cell(row = 3, column =3+(4*k) ,value = 'Markers for the next generation')
		sheet_Result.cell(row = 4,column = 1+(4*k), value= sp[0])
		sheet_Result.cell(row = 4,column = 2+(4*k), value= val[0])	

		m=0
		for l in rawData.loc[sp[0]]:
			if l in ['-','A','H'] or l != voteDye[m]:
				sheet_Result.cell(row = j, column = 3+(4*k), value = rawData[m][0])
				j+=1
			m+=1
		k+=1

		html_candidate+="""<tr><td>"""+str(k)+"""</td><td>"""+str(sp[0])+"""</td><td>"""+str(val[0])+"""</td><td>"""+str(j-4)+"""</td></tr>"""

	html_candidate += """</table></html>"""

		#print (indexMark)
	for i in list(sp_bgScore.index):
		val=list(sp_bgScore.loc[[i]].value)
		indexMark = int(val[0])
		sp_name = list(sp_bgScore.loc[[i]]['sample'])
		label[indexMark-int(min(hist_data))] += str(sp_name[0])+":"+str(val[0])+" "
	
	#color=[50]*20

	#label = []	#for create histogram
	#print(color)
	#print(label)
	#print(rf_bgScore)
	#print(sp_bgScore)
	#print(list(sp_sel['value']))

	### Histrogram ###
	bin_size = 1
	hist = dict(type='histogram',
			x=hist_data,
			xaxis='x1',
			yaxis='y1',
			#histnorm='probability density',
			name= gen.upper() + "_Histrogram",
			legendgroup=group_labels,
			marker=dict(
					cmax= 100, 
					cmin= 0, 
					color= color,
					colorscale= colorscale
					),
			autobinx=False,
			xbins=dict(start=0,
						end=100,
						size=1
						),
			text = label,
			hoverinfo = "y+text"
			)
	#print(hist)
	#print (hist['yaxis'])
	'''### Normal curve ###
	mean, sd = (scipy.stats.norm.fit(hist_data))
	curve_x = [i for i in range(100)]
	#curve_x = [min(hist_data) +	i * (max(hist_data) - min(hist_data))/ 100 for i in range(100)]
	curve_y = scipy.stats.norm.pdf(curve_x, loc=mean, scale=sd)	#prob density
	curve_y *= len(rawData)*2
	curve_y *= bin_size
	#print (curve_y)
	#print (label)
	curve= dict(type='scatter',x=curve_x,
				y=curve_y,
				xaxis='x1',
				yaxis='y1',
				mode='lines',
				showlegend=False,
				hoverinfo = 'none',
				line = dict(
						color = '#F1948A',
						smoothing = 1.3
						)
				)
				
	shadeCurve = py.graph_objs.Scatter(x = curve_x[int(minEspect):],
							y = curve_y[minEspect:],
							fill = 'tozeroy',
							line = dict(width = 0.1,
									color='#F1948A'
								),
							name = gen.upper() + "_Distribution",
							mode = 'lines',
							hoverinfo = 'none',
							)'''
	### Layout ###							
						
	layout = {
		  "autosize": True, 
		  "height": 582, 
		  "title": gen.upper(), 
		  "width": 1202, 
		  "xaxis": {
			"autorange": False, 
			"range": [0,100], 	
			"type": "linear",
			"title" : "Percent of recurrent parent genome",
			"tick0" : 0,
			"dtick" : 10
		  }, 
		  "yaxis": {
			"autorange": True, 
			"range": [0, 1],
			"title" : "Frequency"
		  }
		}

	#fig = ff.create_distplot(hist_data, group_labels, bin_size=.5, curve_type='kde', rug_text = text, histnorm= 'probability')
		### Create ###
	data = [[hist],]
	data = sum(data, [])
	fig = Figure(data=data, layout=layout)
	#py.plotly.image.save_as(fig,'graph.png')
	#py.offline.init_notebook_mode()
	py.offline.plot({"data": data, "layout": layout}, filename='tmp.html', auto_open=False)
	#os.system('rm /home/amikaya/Downloads/graph.png')
	#py.offline.plot(fig, image = 'png', image_filename='graph' , output_type='file', filename='tmp.html', auto_open=False)
	#img = Image('/home/amikaya/Downloads/graph.png')
	#img.anchor(sheet_Result.cell(row = 3,column=5))
	#sheet_Result.add_image(img)
	#os.system('less download.html >> tmp.html')

wb.save(args.file)
#w_retestMarker.close()
os.system('cp '+args.file+' /var/www/html/back_cross')

html_stat = """
<html>
<body>
<p>Generation: """+gen.upper()+"""</p>
<p>Number of plant samples: """+str(laSp_row-fiSp_row)+"""</p>
<p>Minimum expectation of %RPG (min.expect): """+str(minEspect)+"""</p>
<p>Number of plant with %RPG &#8805; min.expect: """+str(plantMoreExpect)+"""</p>
</body>
</html>

"""

html_name=str(args.file).replace('.xlsx','.html')
os.system('cp tmp.html '+html_name)
os.system('cp '+html_name+' /var/www/html/back_cross')

html_download = """
<html>
<body>
<p style="text-align:center;font-family:verdana" > 
Click on the link to download your result </p>
<a href="http://192.168.5.82/back_cross/"""+args.file+"""" download" >
<p style="text-align:center;font-family:arial"> xlsx file </p>
<a href="http://192.168.5.82/back_cross/"""+html_name+"""" download" >
<p style="text-align:center;font-family:arial"> distribution graph </p>
</a>
</body>
</html>"""

f=open('tmp.html','r')
a=f.read()
a=a.replace(': true})',': false})')
f.close()
w=open('tmp.html','w')
w.write(html_stat+'\n')
w.write(a+'\n')
w.write(html_download+'\n')
w.write(html_candidate)
w.close()


##################################################################################################
#######################################  FMAS Manegement  ########################################
##################################################################################################
	
	

	


